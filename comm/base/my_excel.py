#!/usr/bin/env python
# -*- coding: utf-8 -*-
# author: R.
# email: sy.mailbox@foxmail.com
from openpyxl import load_workbook


class FileTools():
    def excel_file(self, filename, sheetname):  # 打开Excel文件并读取数据
        wb = load_workbook(filename)  # 打开指定Excel文件
        sheet = wb[sheetname]  # 打开指定工作表
        total_rows = sheet.max_row  # 获取总行数
        art_data = []  # 新建一个空列表，将读取出来的每行数据存放到列表中
        for x in range(2, total_rows + 1):  # 读取每行数据
            case_data = []  # 组装每行列表数据，形成一个列表集合
            for y in range(3, 8):  # 获取第3列到第7列的数据
                case_data.append(sheet.cell(row=x, column=y).value)  # 将每行单元格数据添加到case_data列表中
            art_data.append(case_data)  # 将每行数据添加到art_data列表中
        return art_data
